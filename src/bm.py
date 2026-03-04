import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import sqlite3
import shutil
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import json
import unicodedata
from src.config import carregar_config
from src.excel import ler_planilha
from src.database import get_conn

CELULAS_BM = {
    'numero_bm': 'F9',
    'endereco_obra': 'D7',
    'titulo_obra': 'A1',
    'data_emissao': 'G9',
    'linha_inicial_locacoes': 15,
    'valor_total': 'C35',
    'revisao': 'G10'
}

COLUNAS_LOCACOES_BM = {
    'descricao': 'A',
    'quantidade': 'B',
    'periodo': 'C',
    'dias': 'D',
    'valor_mensal': 'E',
    'valor_cobrar': 'F',
    'observacoes': 'G'
}

def buscar_ultimo_bm(os):
    conn = get_conn()   
    cursor = conn.cursor()

    cursor.execute('''
        SELECT MAX(numero_bm) FROM boletins WHERE os = ?
    ''', (str(os),))
    
    resultado = cursor.fetchone()[0]
    conn.close()
    
    return resultado if resultado else 0

def buscar_dados_obra(os):
    conn = get_conn()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT titulo_obra, endereco_obra FROM config_os WHERE os = ?
    ''', (str(os),))
    
    resultado = cursor.fetchone()
    conn.close()
    
    if resultado:
        return {'titulo': resultado[0], 'endereco': resultado[1]}
    return None

def normalizar(texto):
    texto = texto.upper()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    return texto

def buscar_valor_mensal(modelo):
    with open('config/precos.json', 'r', encoding='utf-8') as f:
        precos = json.load(f)
    
    modelo_norm = normalizar(modelo)
    
    #primeiro tenta exato
    for chave in precos:
        if normalizar(chave) == modelo_norm:
            return precos[chave]
    
    #depois tenta parcial
    for chave in precos:
        if normalizar(chave) in modelo_norm:
            return precos[chave]
    
    return 0

def criar_bm(os, df_os):
    dados_obra = buscar_dados_obra(os)
    if not dados_obra:
        print(f"OS {os} não cadastrada no banco! Pulando...")
        return None

    numero_bm = buscar_ultimo_bm(os) + 1
    hoje = datetime.now()
    
    config = carregar_config()
    raiz = config['caminhos']['raiz_faturamento']
    pasta_mes = f"{hoje.month:02d}-{hoje.year}"
    pasta_os = Path(raiz) / pasta_mes / f"OS_{os}"
    pasta_os.mkdir(parents=True, exist_ok=True)
    
    nome_arquivo = f"Boletim de Medicao - OS {os} - BM{numero_bm:02d}.xlsx"
    caminho_destino = pasta_os / nome_arquivo
    
    template = Path('templates/Template_BM.xlsx')
    shutil.copy(template, caminho_destino)
    
    wb = load_workbook(caminho_destino)
    ws = wb.active

    ws[CELULAS_BM['titulo_obra']] = dados_obra['titulo']
    ws[CELULAS_BM['endereco_obra']] = dados_obra['endereco']
    ws[CELULAS_BM['numero_bm']] = f"{numero_bm:02d}"
    ws[CELULAS_BM['data_emissao']] = hoje.strftime('%d/%m/%Y')

    linha_atual = CELULAS_BM['linha_inicial_locacoes']
    valor_total = 0
    containers = {}

    for _, row in df_os.iterrows():
        tipo = str(row['TIPO DO VEICULO']).upper()

        if 'CONTAINER' in tipo:
            modelo = row['MODELO']
            if modelo not in containers:
                containers[modelo] = {
                    'quantidade': 0,
                    'inicio': row['INICIO'],
                    'fim': row['FIM'],
                    'dias': 0,
                    'valor_cobrar': 0
                }
            containers[modelo]['quantidade'] += 1
            containers[modelo]['dias'] = row['DIAS']
            containers[modelo]['valor_cobrar'] += row['A COBRAR']

        else:
            descricao = f"{row['Nº FROTA']} - {row['MODELO'].title()} - Placa {row['PLACA/CHASSI']}"
            valor_mensal = buscar_valor_mensal(str(row['TIPO DO VEICULO']))
            inicio = pd.to_datetime(row['INICIO'], dayfirst=True, errors='coerce')
            fim = pd.to_datetime(row['FIM'], dayfirst=True, errors='coerce')
            periodo = f"{inicio.strftime('%d/%m/%Y')} até {fim.strftime('%d/%m/%Y')}"

            ws[f"{COLUNAS_LOCACOES_BM['descricao']}{linha_atual}"] = descricao
            ws[f"{COLUNAS_LOCACOES_BM['quantidade']}{linha_atual}"] = 1
            ws[f"{COLUNAS_LOCACOES_BM['periodo']}{linha_atual}"] = periodo
            ws[f"{COLUNAS_LOCACOES_BM['dias']}{linha_atual}"] = row['DIAS']
            ws[f"{COLUNAS_LOCACOES_BM['valor_mensal']}{linha_atual}"] = valor_mensal
            ws[f"{COLUNAS_LOCACOES_BM['valor_cobrar']}{linha_atual}"] = row['A COBRAR']

            valor_total += float(row['A COBRAR'] or 0)
            linha_atual += 1

    for modelo, dados in containers.items():
        valor_mensal = buscar_valor_mensal(str(modelo))
        inicio = pd.to_datetime(dados['inicio'], dayfirst=True, errors='coerce')
        fim = pd.to_datetime(dados['fim'], dayfirst=True, errors='coerce')
        periodo = f"{inicio.strftime('%d/%m/%Y')} até {fim.strftime('%d/%m/%Y')}"

        ws[f"{COLUNAS_LOCACOES_BM['descricao']}{linha_atual}"] = modelo.title()
        ws[f"{COLUNAS_LOCACOES_BM['quantidade']}{linha_atual}"] = dados['quantidade']
        ws[f"{COLUNAS_LOCACOES_BM['periodo']}{linha_atual}"] = periodo
        ws[f"{COLUNAS_LOCACOES_BM['dias']}{linha_atual}"] = dados['dias']
        ws[f"{COLUNAS_LOCACOES_BM['valor_mensal']}{linha_atual}"] = valor_mensal
        ws[f"{COLUNAS_LOCACOES_BM['valor_cobrar']}{linha_atual}"] = dados['valor_cobrar']

        valor_total += float(dados['valor_cobrar'] or 0)
        linha_atual += 1

    ws[CELULAS_BM['valor_total']] = valor_total

    wb.save(caminho_destino)

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO boletins (os, numero_bm, status, data_criacao, caminho_xlsx)
        VALUES (?, ?, 'criado', ?, ?)
    ''', (str(os), numero_bm, datetime.now().isoformat(), str(caminho_destino)))
    conn.commit()
    conn.close()

    print(f"BM criado: {nome_arquivo}")
    return str(caminho_destino)

def gerar_todos_bms():
    hoje = datetime.now()
    df, oss = ler_planilha(hoje.month - 1, hoje.year)
    
    print(f"\nGerando BMs para {len(oss)} OSs...\n")
    
    for os in oss:
        df_os = df[df['OS'] == os]
        criar_bm(os, df_os)
    
    print("\nTodos os BMs gerados!")

if __name__ == "__main__":
    gerar_todos_bms()