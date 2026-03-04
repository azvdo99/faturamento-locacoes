import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from src.database import get_conn

def listar_faturas_criadas():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, numero_fat, os, caminho_xlsx FROM faturas 
        WHERE status = 'criada'
        ORDER BY numero_fat
    ''')
    faturas = cursor.fetchall()
    conn.close()
    return [{'id': f[0], 'numero_fat': f[1], 'os': f[2], 'caminho_xlsx': f[3]} for f in faturas]

def ler_locacoes_fatura(caminho_xlsx):
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    locacoes = []
    linha = 18
    while True:
        descricao = ws[f'B{linha}'].value
        if not descricao or not str(descricao).startswith('Locação'):
            break
        locacoes.append({'linha': linha, 'descricao': descricao})
        linha += 5
    wb.close()
    return locacoes

def salvar_pedidos(caminho_xlsx, pedidos):
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    for linha, pedido in pedidos:
        ws[f'H{linha}'] = pedido
    wb.save(caminho_xlsx)
    wb.close()

def atualizar_status_fatura(fatura_id):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('UPDATE faturas SET status = ? WHERE id = ?', ('pronta', fatura_id))
    conn.commit()
    conn.close()

def main():
    faturas = listar_faturas_criadas()
    
    if not faturas:
        print("\nNenhuma fatura aguardando pedidos!")
        return
    
    print("\n" + "=" * 50)
    print("FATURAS AGUARDANDO PEDIDOS")
    print("=" * 50 + "\n")
    
    for i, fat in enumerate(faturas, 1):
        print(f"{i}. FAT {fat['numero_fat']} - OS {fat['os']}")
    
    try:
        escolha = int(input("\nEscolha uma fatura: "))
        if escolha < 1 or escolha > len(faturas):
            print("Opção inválida!")
            return
    except ValueError:
        print("Digite um número válido!")
        return
    
    fatura = faturas[escolha - 1]
    locacoes = ler_locacoes_fatura(fatura['caminho_xlsx'])
    
    print(f"\nFAT {fatura['numero_fat']} - OS {fatura['os']}")
    print(f"{len(locacoes)} locações encontradas\n")
    
    pedidos = []
    for loc in locacoes:
        print(f"→ {loc['descricao']}")
        pedido = input("  Pedido (ou Enter para SEM PC): ").strip()
        
        if not pedido:
            valor_final = 'SEM PC'
        elif not pedido.upper().startswith('PC'):
            valor_final = f"PC {pedido}"
        else:
            valor_final = pedido
            
        pedidos.append((loc['linha'], valor_final))
    
    salvar_pedidos(fatura['caminho_xlsx'], pedidos)
    atualizar_status_fatura(fatura['id'])
    
    print(f"\nPedidos salvos! FAT {fatura['numero_fat']} pronta.")

if __name__ == "__main__":
    main()