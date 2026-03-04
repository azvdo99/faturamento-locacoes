import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import smtplib
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from email.utils import make_msgid
from src.database import get_conn
from src.config import carregar_config


def get_saudacao():
    hora = datetime.now().hour
    if hora < 12:
        return "bom dia"
    elif hora < 18:
        return "boa tarde"
    else:
        return "boa noite"

def adicionar_anexo(msg, caminho_pdf):
    with open(caminho_pdf, 'rb') as f:
        mime = MIMEBase('application', 'octet-stream')
        mime.set_payload(f.read())
    encoders.encode_base64(mime)
    mime.add_header('Content-Disposition', 'attachment', filename=Path(caminho_pdf).name)
    msg.attach(mime)

def verificar_sem_pc(caminho_xlsx):
    wb = load_workbook(caminho_xlsx, data_only=True)
    ws = wb.active
    
    sem_pc = []
    linha = 16
    
    while True:
        descricao = ws[f'B{linha}'].value
        if not descricao or not str(descricao).startswith('Locação'):
            break
        
        pedido = ws[f'H{linha}'].value
        if not pedido or str(pedido).strip() in ['', 'SEM PC']:
            sem_pc.append(descricao)
        
        linha += 5
    
    wb.close()
    return sem_pc

def enviar_bm(bm_id):
    config = carregar_config()
    
    with open('config/emails_obras.json', 'r', encoding='utf-8') as f:
        emails_config = json.load(f)
    
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT os, numero_bm, caminho_pdf FROM boletins WHERE id = ?
    ''', (bm_id,))
    bm = cursor.fetchone()
    conn.close()
    
    if not bm:
        print(f"BM {bm_id} não encontrado!")
        return False
    
    os_str, numero_bm, caminho_pdf = bm
    
    if not caminho_pdf:
        print(f"PDF do BM {bm_id} não gerado ainda!")
        return False
    
    if os_str not in emails_config['destinatarios']:
        print(f"Emails da OS {os_str} não cadastrados!")
        return False
    
    smtp = config['smtp']
    sender = smtp['usuario']
    
    destinatarios = emails_config['destinatarios'][os_str]
    copias = emails_config.get('copia', []).copy()
    if os_str in emails_config.get('copias_por_os', {}):
        copias.append(emails_config['copias_por_os'][os_str])
    
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = ', '.join(destinatarios)
    if copias:
        msg['Cc'] = ', '.join(copias)
    msg['Subject'] = f"Boletim de Medição N°{numero_bm:02d} - OS {os_str}"
    
    GIF = "https://via.placeholder.com/600x200?text=Logo+Exemplo"
    corpo = f"""Prezados, {get_saudacao()}!<br><br>
Estou encaminhando em anexo o Boletim de Medição N°{numero_bm:02d}, referente a locação dos veículos/equipamentos da OS {os_str}.<br><br>
Este documento foi elaborado para sua revisão e aprovação antes de prosseguirmos com o faturamento.<br><br>
Por favor, revisem o boletim atentamente e confirmem se está de acordo com os serviços prestados.<br><br>
Caso haja qualquer questionamento ou necessidade de esclarecimento, não hesite em entrar em contato.<br><br>
Fico à disposição!<br><br>
<img src="{GIF}" alt="Locadora Exemplo" style="max-width:100%;height:auto;">"""
    
    msg.attach(MIMEText(corpo, 'html', 'utf-8'))
    adicionar_anexo(msg, caminho_pdf)

    message_id = make_msgid()
    msg['Message-ID'] = message_id
        
    try:
        todos = destinatarios + copias
        with smtplib.SMTP(smtp['host'], smtp['porta']) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(sender, smtp['senha'])
            server.sendmail(sender, todos, msg.as_string())
        
        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE boletins SET status = 'enviado', data_envio = ?, email_message_id = ? WHERE id = ?
        ''', (datetime.now().isoformat(), message_id, bm_id))
        conn.commit()
        conn.close()
        
        print(f"BM {numero_bm:02d} - OS {os_str} enviado!")
        return True
        
    except Exception as e:
        print(f"Erro ao enviar BM OS {os_str}: {e}")
        return False

def enviar_todos_bms():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id FROM boletins WHERE status = 'criado' AND caminho_pdf IS NOT NULL
    ''')
    bms = cursor.fetchall()
    conn.close()
    
    if not bms:
        print("Nenhum BM pronto pra enviar!")
        return
    
    print(f"\nEnviando {len(bms)} BMs...\n")
    
    for (bm_id,) in bms:
        enviar_bm(bm_id)

def enviar_fatura(fatura_id):
    config = carregar_config()
    
    with open('config/emails_obras.json', 'r', encoding='utf-8') as f:
        emails_config = json.load(f)
    
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT f.os, f.numero_fat, f.caminho_pdf, f.caminho_xlsx, b.email_message_id
        FROM faturas f
        JOIN boletins b ON f.bm_id = b.id
        WHERE f.id = ?
    ''', (fatura_id,))
    fatura = cursor.fetchone()
    conn.close()
    
    if not fatura:
        print(f"Fatura {fatura_id} não encontrada!")
        return False
    
    os_str, numero_fat, caminho_pdf, caminho_xlsx, message_id_bm = fatura
    
    if not caminho_pdf:
        print(f"PDF da fatura {fatura_id} não gerado ainda!")
        return False
    
    sem_pc = verificar_sem_pc(caminho_xlsx)
    eng_rc = emails_config.get('eng_rc', {}).get(os_str, '')
    
    if sem_pc and not eng_rc:
        print(f"OS {os_str} tem locações sem PC e sem eng_rc cadastrado. Cadastre no emails_obras.json!")
        return False
    
    if os_str not in emails_config['destinatarios']:
        print(f"Emails da OS {os_str} não cadastrados!")
        return False
    
    smtp = config['smtp']
    sender = smtp['usuario']
    
    destinatarios = emails_config['destinatarios'][os_str]
    copias = emails_config.get('copia', []).copy()
    if os_str in emails_config.get('copias_por_os', {}):
        copias.append(emails_config['copias_por_os'][os_str])
    
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = ', '.join(destinatarios)
    if copias:
        msg['Cc'] = ', '.join(copias)
    msg['Subject'] = f"Fatura N°{numero_fat} - OS {os_str}"
    
    if message_id_bm:
        msg['Subject'] = "Re: " + msg['Subject']
        msg['In-Reply-To'] = message_id_bm
        msg['References'] = message_id_bm
    
    GIF = "https://enind.com.br/wp-content/uploads/2024/03/Automacao-ENIND-4-1-1.gif"
    
    if sem_pc and eng_rc:
        corpo = f"""Prezados, {get_saudacao()}!<br><br>
Segue anexo fatura emitida conforme aprovação.<br><br>
@{eng_rc} Por gentileza emitir RC para:<br>"""
        for loc in sem_pc:
            corpo += f"- {loc}<br>"
        corpo += "<br>Se possível já efetuar a RC com previsão.<br><br>"
    else:
        corpo = f"""Prezados, {get_saudacao()}!<br><br>
Conforme aprovação, segue em anexo a fatura N°{numero_fat} referente à OS {os_str}.<br><br>
Fico à disposição para qualquer esclarecimento.<br><br>"""
    
    corpo += f'<img src="{GIF}" alt="Locadora Exemplo" style="max-width:100%;height:auto;">'
    
    msg.attach(MIMEText(corpo, 'html', 'utf-8'))
    adicionar_anexo(msg, caminho_pdf)
    
    try:
        todos = destinatarios + copias
        with smtplib.SMTP(smtp['host'], smtp['porta']) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(sender, smtp['senha'])
            server.sendmail(sender, todos, msg.as_string())
        
        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE faturas SET status = 'enviada', data_envio = ? WHERE id = ?
        ''', (datetime.now().isoformat(), fatura_id))
        conn.commit()
        conn.close()
        
        print(f"Fatura {numero_fat} - OS {os_str} enviada!")
        return True
        
    except Exception as e:
        print(f"Erro ao enviar fatura OS {os_str}: {e}")
        return False

def enviar_todas_faturas():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id FROM faturas WHERE status = 'pronta' AND caminho_pdf IS NOT NULL
    ''')
    faturas = cursor.fetchall()
    conn.close()
    
    if not faturas:
        print("Nenhuma fatura pronta pra enviar!")
        return
    
    print(f"\nEnviando {len(faturas)} faturas...\n")
    
    for (fatura_id,) in faturas:
        enviar_fatura(fatura_id)    