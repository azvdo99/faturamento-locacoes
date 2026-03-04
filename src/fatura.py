import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import shutil
import calendar
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from src.database import get_conn
from src.config import carregar_config
from src.excel import ler_planilha
from src.bm import buscar_valor_mensal

CELULAS_FAT = {
    'numero_fat': 'H4',
    'valor_total': 'I4',
    'data_emissao': 'H6',
    'data_vencimento': 'I6',
    'os_bm': 'B16',
    'valor_liquido': 'I63',
    'linha_inicial_locacoes': 18
}

def buscar_ultimo_fat():
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('SELECT MAX(numero_fat) FROM faturas')
    resultado = cursor.fetchone()[0]
    conn.close()
    return resultado if resultado else 0

def ultimo_dia_util_mes():
    hoje = datetime.now()
    ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
    data_fim = datetime(hoje.year, hoje.month, ultimo_dia)
    while data_fim.weekday() > 4:
        data_fim = data_fim.replace(day=data_fim.day - 1)
    return data_fim

def buscar_bm_aprovado(os):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, numero_bm FROM boletins 
        WHERE os = ? AND status = 'aprovado'
        ORDER BY numero_bm DESC LIMIT 1
    ''', (str(os),))
    resultado = cursor.fetchone()
    conn.close()
    if resultado:
        return {'id': resultado[0], 'numero_bm': resultado[1]}
    return None

def criar_fatura(os, numero_fat, df_os):
    bm = buscar_bm_aprovado(os)
    if not bm:
        print(f"Nenhum BM aprovado para OS {os}! Pulando...")
        return None

    hoje = datetime.now()
    nome_arquivo = f"FAT {numero_fat} - OS {os} - Fatura de Cobrança - Locadora Exemplo.xlsx"

    config = carregar_config()
    raiz = config['caminhos']['raiz_faturamento']
    pasta_mes = f"{hoje.month:02d}-{hoje.year}"
    pasta_os = Path(raiz) / pasta_mes / f"OS_{os}"
    pasta_os.mkdir(parents=True, exist_ok=True)

    caminho_destino = pasta_os / nome_arquivo

    template = Path('templates/Template_FAT.xlsx')
    shutil.copy(template, caminho_destino)

    wb = load_workbook(caminho_destino)
    ws = wb.active

    ws[CELULAS_FAT['numero_fat']] = f"{numero_fat}/{hoje.year}"
    ws[CELULAS_FAT['data_emissao']] = hoje
    ws[CELULAS_FAT['data_vencimento']] = ultimo_dia_util_mes()
    ws[CELULAS_FAT['os_bm']] = f"OS {os} - BM{bm['numero_bm']:02d}"

    linha_atual = CELULAS_FAT['linha_inicial_locacoes']
    containers = {}

    for _, row in df_os.iterrows():
        tipo = str(row['TIPO DO VEICULO']).upper()

        if 'CONTAINER' in tipo:
            modelo = row['TIPO DO VEICULO']
            if modelo not in containers:
                containers[modelo] = {
                    'inicio': row['INICIO'],
                    'fim': row['FIM'],
                    'dias': row['DIAS'],
                    'valor_cobrar': 0,
                    'valor_mensal': buscar_valor_mensal(str(row['TIPO DO VEICULO']))
                }
            containers[modelo]['valor_cobrar'] += row['A COBRAR']

        else:
            valor_mensal = buscar_valor_mensal(str(row['TIPO DO VEICULO']))
            inicio = pd.to_datetime(row['INICIO'], dayfirst=True, errors='coerce')
            fim = pd.to_datetime(row['FIM'], dayfirst=True, errors='coerce')
            placa = str(row['PLACA/CHASSI']) if pd.notna(row['PLACA/CHASSI']) else 'S/Placa'

            ws[f"B{linha_atual}"] = f"Locação de {row['TIPO DO VEICULO'].title()} - {row['MODELO'].title()} - Placa {placa}"
            ws[f"H{linha_atual}"] = ""
            ws[f"I{linha_atual}"] = row['A COBRAR']
            ws[f"B{linha_atual + 1}"] = f"Valor Mensal - R${valor_mensal:,.2f}"
            ws[f"B{linha_atual + 2}"] = f"Período: {inicio.strftime('%d/%m/%Y')} até {fim.strftime('%d/%m/%Y')}"
            ws[f"B{linha_atual + 3}"] = f"Dias cobrados: {row['DIAS']}"

            linha_atual += 5

    for modelo, dados in containers.items():
        inicio = pd.to_datetime(dados['inicio'], dayfirst=True, errors='coerce')
        fim = pd.to_datetime(dados['fim'], dayfirst=True, errors='coerce')

        ws[f"B{linha_atual}"] = f"Locação de {modelo.title()}"
        ws[f"H{linha_atual}"] = ""
        ws[f"I{linha_atual}"] = dados['valor_cobrar']
        ws[f"B{linha_atual + 1}"] = f"Valor Mensal - R${dados['valor_mensal']:,.2f}"
        ws[f"B{linha_atual + 2}"] = f"Período: {inicio.strftime('%d/%m/%Y')} até {fim.strftime('%d/%m/%Y')}"
        ws[f"B{linha_atual + 3}"] = f"Dias cobrados: {dados['dias']}"

        linha_atual += 5

    valor_total = sum(
        float(row['A COBRAR'] or 0) for _, row in df_os.iterrows()
    )
    
    ws[CELULAS_FAT['valor_total']] = valor_total
    ws[CELULAS_FAT['valor_liquido']] = valor_total

    wb.save(caminho_destino)

    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO faturas (numero_fat, os, bm_id, status, data_criacao, data_vencimento, caminho_xlsx)
        VALUES (?, ?, ?, 'criada', ?, ?, ?)
    ''', (numero_fat, str(os), bm['id'], datetime.now().isoformat(), ultimo_dia_util_mes().isoformat(), str(caminho_destino)))
    conn.commit()
    conn.close()

    print(f"Fatura criada: {nome_arquivo}")
    return str(caminho_destino)

def gerar_todas_faturas(numero_fat_inicial):
    conn = get_conn()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DISTINCT os FROM boletins WHERE status = 'aprovado'
    ''')
    oss_aprovadas = [row[0] for row in cursor.fetchall()]
    conn.close()

    if not oss_aprovadas:
        print("Nenhum BM aprovado encontrado!")
        return

    config = carregar_config()
    hoje = datetime.now()
    df, _ = ler_planilha(hoje.month - 1, hoje.year)

    print(f"\nGerando faturas para {len(oss_aprovadas)} OSs...\n")

    numero_fat_atual = numero_fat_inicial
    for os in oss_aprovadas:
        df_os = df[df['OS'] == float(os)]
        criar_fatura(os, numero_fat_atual, df_os)
        numero_fat_atual += 1

    print("\nTodas as faturas geradas!")

if __name__ == "__main__":
    ultimo = input("Último número de fatura usado (ou Enter para buscar): ").strip()
    numero_inicial = int(ultimo) + 1 if ultimo else buscar_ultimo_fat() + 1
    print(f"\nIniciando em: {numero_inicial}")
    gerar_todas_faturas(numero_inicial)